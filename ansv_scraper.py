#!/usr/bin/env python3
"""
Script de PRODUCCIÓN (Scraper + API)
- Scraper: Se ejecuta diariamente a las 6 AM mediante cron (llamado con "scrape")
- API: Se ejecuta con Flask para servir los datos como JSON.
"""

from selenium import webdriver
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import os
import time
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
import shutil
import logging
import sys
import json
from flask import Flask, jsonify, request, abort

# --- Configuración de Logging ---
log_dir = "/app/logs" # Usar ruta absoluta
if not os.path.exists(log_dir):
    os.makedirs(log_dir)

log_file = os.path.join(log_dir, f"ansv_{datetime.now().strftime('%Y%m')}.log")
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_file),
        logging.StreamHandler(sys.stdout)
    ]
)

# --- Configuración de la Aplicación ---
DATA_DIR = "/app/data" # Directorio de volúmen de datos
app = Flask(__name__)

# ======================================================================
# LÓGICA DEL SCRAPER (Descarga de archivos)
# ======================================================================

def descargar_excel_temporal(url):
    """
    Descarga el archivo Excel a una carpeta temporal
    """
    carpeta_temp = os.path.join(DATA_DIR, "temp_descargas")
    
    if os.path.exists(carpeta_temp):
        shutil.rmtree(carpeta_temp)
    os.makedirs(carpeta_temp)
    
    firefox_options = Options()
    firefox_options.add_argument("--headless")
    firefox_options.add_argument("--no-sandbox")
    firefox_options.add_argument("--disable-dev-shm-usage")
    firefox_options.add_argument("--disable-gpu")
    firefox_options.add_argument("--window-size=1920,1080")
    
    firefox_options.set_preference("browser.download.folderList", 2)
    firefox_options.set_preference("browser.download.dir", carpeta_temp)
    firefox_options.set_preference("browser.download.useDownloadDir", True)
    firefox_options.set_preference("browser.download.prompt.for.download", False)
    firefox_options.set_preference("browser.helperApps.neverAsk.saveToDisk", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/vnd.ms-excel,application/octet-stream")
    
    try:
        service = Service(executable_path="/usr/local/bin/geckodriver")
        driver = webdriver.Firefox(service=service, options=firefox_options)
        
        logging.info("Accediendo a la página de ANSV...")
        driver.get(url)
        time.sleep(3)
        
        iframe = WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.TAG_NAME, "iframe"))
        )
        driver.switch_to.frame(iframe)
        logging.info("Iframe encontrado, cambiando contexto...")
        time.sleep(2)
        
        selectores_posibles = [
            "//div[contains(@class, 'dx-datagrid-export-button')]",
            "//div[@aria-label='export-excel-button']",
            "//div[@title='Exportar todo']",
        ]
        
        boton = None
        for selector in selectores_posibles:
            try:
                boton = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, selector))
                )
                break
            except:
                continue
        
        if not boton:
            logging.error("No se encontró el botón de descarga")
            driver.quit()
            return None
        
        logging.info("Botón de descarga encontrado, haciendo clic...")
        try:
            boton.click()
        except:
            driver.execute_script("arguments[0].click();", boton)
        
        tiempo_espera = 60
        tiempo_transcurrido = 0
        
        while tiempo_transcurrido < tiempo_espera:
            archivos = os.listdir(carpeta_temp)
            archivos_excel = [f for f in archivos if f.endswith(('.xlsx', '.xls')) and not f.endswith('.crdownload')]
            
            if archivos_excel:
                archivo_descargado = os.path.join(carpeta_temp, archivos_excel[0])
                logging.info(f"Archivo descargado exitosamente: {archivos_excel[0]}")
                driver.quit()
                return archivo_descargado
            
            time.sleep(1)
            tiempo_transcurrido += 1
        
        logging.error("Tiempo de espera agotado para la descarga")
        driver.quit()
        return None
        
    except Exception as e:
        logging.error(f"Error durante la descarga: {str(e)}")
        if 'driver' in locals():
            driver.quit()
        return None

def extraer_fecha_del_nombre(nombre_archivo):
    """
    Extrae la fecha del nombre del archivo descargado
    """
    try:
        nombre_base = os.path.basename(nombre_archivo)
        fecha_parte = nombre_base.split('-')[0]
        
        if len(fecha_parte) >= 8:
            año = fecha_parte[0:4]
            mes = fecha_parte[4:6]
            dia = fecha_parte[6:8]
            
            fecha_formateada = f"{año}-{mes}-{dia}"
            logging.info(f"Fecha extraída del archivo: {fecha_formateada}")
            return fecha_formateada
        else:
            return None
    except Exception as e:
        logging.warning(f"No se pudo extraer fecha del nombre: {str(e)}")
        return None

def obtener_nombre_mes(numero_mes):
    """
    Convierte número de mes a nombre en español
    """
    meses = {
        1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
        5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
        9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
    }
    return meses.get(numero_mes, "Desconocido")

def agregar_hoja_a_excel(archivo_origen, archivo_destino, nombre_hoja):
    """
    Agrega el contenido del archivo_origen como una nueva hoja en archivo_destino
    """
    try:
        df = pd.read_excel(archivo_origen)
        
        if not os.path.exists(archivo_destino):
            logging.info(f"Creando nuevo archivo: {os.path.basename(archivo_destino)}")
            df.to_excel(archivo_destino, sheet_name=nombre_hoja, index=False)
        else:
            logging.info(f"Agregando hoja '{nombre_hoja}' al archivo existente")
            with pd.ExcelWriter(archivo_destino, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name=nombre_hoja, index=False)
        
        return True
    except Exception as e:
        logging.error(f"Error al agregar hoja: {str(e)}")
        return False

def ejecutar_descarga_diaria():
    """
    Ejecuta la descarga diaria del archivo Excel de ANSV
    """
    url = "https://fotodeteccion.ansv.gov.co/ubicaciones-aprobadas.html"
    
    ahora = datetime.now()
    año_actual = ahora.year
    mes_actual = ahora.month
    dia_actual = ahora.day
    
    carpeta_año = os.path.join(DATA_DIR, str(año_actual))
    if not os.path.exists(carpeta_año):
        os.makedirs(carpeta_año)
        logging.info(f"Carpeta '{carpeta_año}' creada")
    
    nombre_mes = obtener_nombre_mes(mes_actual)
    archivo_mes = os.path.join(carpeta_año, f"{nombre_mes}.xlsx")
    
    logging.info("=" * 70)
    logging.info("INICIO DE DESCARGA DIARIA - ANSV FOTODETECCIÓN")
    logging.info("=" * 70)
    logging.info(f"Fecha: {ahora.strftime('%Y-%m-%d %H:%M:%S')}")
    logging.info(f"Archivo destino: {archivo_mes}")
    
    logging.info("Iniciando descarga desde ANSV...")
    archivo_temporal = descargar_excel_temporal(url)
    
    if archivo_temporal:
        fecha_extraida = extraer_fecha_del_nombre(archivo_temporal)
        
        if fecha_extraida:
            nombre_hoja = fecha_extraida
        else:
            nombre_hoja = f"{año_actual}-{mes_actual:02d}-{dia_actual:02d}"
            logging.warning(f"Usando fecha actual como fallback: {nombre_hoja}")
        
        logging.info("Procesando y agregando hoja al archivo mensual...")
        if agregar_hoja_a_excel(archivo_temporal, archivo_mes, nombre_hoja):
            logging.info(f"✓ Hoja '{nombre_hoja}' agregada exitosamente")
            
            try:
                wb = load_workbook(archivo_mes)
                logging.info(f"El archivo ahora tiene {len(wb.sheetnames)} hoja(s)")
                logging.info(f"Hojas: {', '.join(wb.sheetnames)}")
                wb.close()
            except:
                pass
            
            logging.info("=" * 70)
            logging.info("✓ DESCARGA COMPLETADA EXITOSAMENTE")
            logging.info("=" * 70)
            
            if os.path.exists(os.path.join(DATA_DIR, "temp_descargas")):
                shutil.rmtree(os.path.join(DATA_DIR, "temp_descargas"))
            
            return True
        else:
            logging.error("Error al agregar la hoja al archivo")
            return False
    else:
        logging.error("No se pudo descargar el archivo")
        logging.info("=" * 70)
        logging.info("✗ DESCARGA FALLIDA")
        logging.info("=" * 70)
        return False

# ======================================================================
# LÓGICA DE LA API (Servidor de datos)
# ======================================================================

def get_excel_data(sheet_name):
    """
    Función auxiliar para leer una hoja específica de un archivo Excel
    """
    try:
        # Determinar el archivo basado en la fecha
        fecha = datetime.strptime(sheet_name, '%Y-%m-%d')
        año = fecha.year
        nombre_mes = obtener_nombre_mes(fecha.month)
        archivo_path = os.path.join(DATA_DIR, str(año), f"{nombre_mes}.xlsx")
        
        if not os.path.exists(archivo_path):
            return None, "Archivo del mes no encontrado"
            
        # Leer la hoja específica
        df = pd.read_excel(archivo_path, sheet_name=sheet_name)
        # Convertir a JSON (orient="records" crea una lista de objetos)
        result_json = df.to_json(orient="records")
        return json.loads(result_json), None
        
    except ValueError:
        return None, "Formato de fecha incorrecto. Usar YYYY-MM-DD"
    except FileNotFoundError:
        return None, "Archivo del mes no encontrado"
    except Exception as e:
        if "No sheet named" in str(e):
             return None, f"Datos para la fecha '{sheet_name}' no encontrados en el archivo"
        logging.error(f"Error al leer Excel: {str(e)}")
        return None, str(e)

@app.route('/')
def index():
    return jsonify({
        "servicio": "API de Scraper ANSV",
        "estado": "en_linea",
        "endpoints": {
            "/api/datos/hoy": "Obtiene los datos de la última descarga (hoy)",
            "/api/datos/fecha/YYYY-MM-DD": "Obtiene los datos para una fecha específica"
        }
    })

@app.route('/api/datos/hoy', methods=['GET'])
def get_datos_hoy():
    # Usar la zona horaria correcta (UTC-5)
    hoy_str = (datetime.now() - pd.Timedelta(hours=5)).strftime('%Y-%m-%d')
    
    datos, error = get_excel_data(hoy_str)
    
    if error:
        return jsonify({"error": error}), 404
    
    return jsonify({
        "fecha": hoy_str,
        "registros": len(datos),
        "datos": datos
    })

@app.route('/api/datos/fecha/<string:fecha>', methods=['GET'])
def get_datos_por_fecha(fecha):
    datos, error = get_excel_data(fecha)
    
    if error:
        return jsonify({"error": error}), 404
    
    return jsonify({
        "fecha": fecha,
        "registros": len(datos),
        "datos": datos
    })

# Este es el punto de entrada para CRON
if __name__ == "__main__":
    # Si el script se llama con "python ansv_scraper.py scrape"
    if len(sys.argv) > 1 and sys.argv[1] == 'scrape':
        try:
            resultado = ejecutar_descarga_diaria()
            sys.exit(0 if resultado else 1)
        except Exception as e:
            logging.error(f"Error crítico en scraper: {str(e)}")
            sys.exit(1)
    else:
        # Este bloque NO se usa cuando se inicia con "flask run"
        # 'flask run' detecta la variable 'app' automáticamente.
        logging.info("Este script debe iniciarse con 'flask run' para la API.")
        logging.info("O con 'python ansv_scraper.py scrape' para el scraper.")
        pass
